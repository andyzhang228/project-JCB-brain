from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os


def write_to_excel(
    template_path, extracted_content, output_path, target_table_name, interface_or_file
):
    # ヘッダーマッピング辞書の作成 - ファイルフォーマット用
    if interface_or_file == "file":
        table_header_map = {
            "#": "#",
            f"FIELD NAME\n(項 目 名)": "項目名",
            f"SYMBOL NAME\n(シ ン ボ ル 名)": "シンボル名",
            f"ATTRIBUTES\n(属性)": "属性",
            f"NUMBER OF DIGITS\n(桁数)": "桁数",
            f"LENGTH\n(バイト数)": "バイト数",
            "OFFSET": "OFFSET",
            "備考": "説明",
        }

    elif interface_or_file == "interface":
        table_header_map = {
            "#": "No",
            f"FIELD NAME\n(項 目 名)": "項目名",
            f"SYMBOL NAME\n(シ ン ボ ル 名)": "項目ＩＤ",
            f"ATTRIBUTES\n(属性)": "タイプ",
            f"NUMBER OF DIGITS\n(桁数)": "Byte2",
            f"LENGTH\n(バイト数)": "Byte",
            f"OFFSET\n(オフセット)": "OFFSET",
            "備考": "処理・備考",
        }

    keys_list = list(table_header_map.keys())
    # print(keys_list)
    col_width = {key: 0 for key in keys_list}
    # print(col_width)
    new_table_header = list(table_header_map.keys())
    # print(new_table_header)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    if os.path.exists(output_path):
        wb_template = load_workbook(output_path)
    else:
        wb_template = load_workbook(template_path)
    # 重複書き込みを防ぎ、データを更新するため、シートが存在する場合は削除して再作成
    if target_table_name in wb_template.sheetnames:
        del wb_template[target_table_name]

    # 新しいシートを作成
    new_sheet = wb_template.create_sheet(title=target_table_name)

    # ヘッダーを書き込み
    for i in range(len(new_table_header)):
        new_sheet.cell(row=1, column=i + 2, value=new_table_header[i])
        # print(new_table_header[i])
        # table design
        new_sheet.cell(row=1, column=i + 2).fill = PatternFill(
            start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"
        )
        new_sheet.cell(row=1, column=i + 2).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        new_sheet.cell(row=1, column=i + 2).border = thin_border
        # 列の幅を更新
        col_width[new_table_header[i]] = len(new_table_header[i])

    # HEADERの高さを設定
    new_sheet.row_dimensions[1].height = 35

    for key, value in table_header_map.items():
        for i in range(len(extracted_content[value])):
            new_sheet.cell(
                row=i + 2,
                column=keys_list.index(key) + 2,
                value=extracted_content[value][i],
            )
            # design
            new_sheet.cell(row=i + 2, column=keys_list.index(key) + 2).alignment = (
                Alignment(horizontal="left", wrap_text=True)
            )
            new_sheet.cell(row=i + 2, column=keys_list.index(key) + 2).border = (
                thin_border
            )
            # 列幅を更新
            str_value = str(extracted_content[value][i])
            if len(str_value) > col_width[key]:
                # print(col_width)
                col_width[key] = len(str_value) * 1.2

    for key, value in col_width.items():
        column_letter = get_column_letter(keys_list.index(key) + 2)
        width = max(10, value * 1.5)
        new_sheet.column_dimensions[column_letter].width = width

    wb_template.save(output_path)
